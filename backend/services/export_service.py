import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from models.evaluation import Evaluation
from models.agent import Agent
from models.monthly_evaluation import MonthlyEvaluation
import json
import io

MESES = ['','Enero','Febrero','Marzo','Abril','Mayo','Junio',
         'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

AZUL   = PatternFill("solid", fgColor="1D4ED8")
VERDE  = PatternFill("solid", fgColor="D1FAE5")
ROJO   = PatternFill("solid", fgColor="FEE2E2")
AMARILLO = PatternFill("solid", fgColor="FEF9C3")
GRIS   = PatternFill("solid", fgColor="F3F4F6")
BOLD_WHITE = Font(bold=True, color="FFFFFF")
BOLD_DARK  = Font(bold=True, color="1F2937")

def _cabecera(ws, texto, celdas, fila):
    ws.merge_cells(celdas)
    ws[f'A{fila}'] = texto
    ws[f'A{fila}'].font = Font(bold=True, size=14, color="1D4ED8")
    ws[f'A{fila}'].alignment = Alignment(horizontal='center')

def _fila_header(ws, headers, fila):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=fila, column=col, value=h)
        cell.fill = AZUL
        cell.font = BOLD_WHITE
        cell.alignment = Alignment(horizontal='center')

def exportar_reporte_excel(agent_id: int, mes: int, anio: int, db: Session) -> bytes:
    agent = db.query(Agent).filter(Agent.id == agent_id).first()
    evaluations = db.query(Evaluation).filter(
        Evaluation.agent_id == agent_id,
        Evaluation.mes == mes,
        Evaluation.anio == anio
    ).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Agente"

    _cabecera(ws, 'Sistema de Supervision - Reporte por Agente', 'A1:F1', 1)

    ws['A3'] = 'Agente:';  ws['A3'].font = BOLD_DARK; ws['B3'] = agent.nombre if agent else 'N/A'
    ws['A4'] = 'Codigo:';  ws['A4'].font = BOLD_DARK; ws['B4'] = agent.codigo if agent else 'N/A'
    ws['A5'] = 'Area:';    ws['A5'].font = BOLD_DARK; ws['B5'] = agent.area if agent else 'N/A'
    ws['A6'] = 'Periodo:'; ws['A6'].font = BOLD_DARK; ws['B6'] = f"{MESES[mes]} {anio}"

    promedio = round(sum(e.puntuacion for e in evaluations) / len(evaluations)) if evaluations else 0
    ws['D3'] = 'Total conversaciones:'; ws['D3'].font = BOLD_DARK; ws['E3'] = len(evaluations)
    ws['D4'] = 'Puntuacion promedio:';  ws['D4'].font = BOLD_DARK; ws['E4'] = f"{promedio}%"

    _fila_header(ws, ['#', 'Conversacion', 'Items cumplidos', 'Items total', 'Puntuacion', 'Estado'], 8)

    for i, ev in enumerate(evaluations, 1):
        row = 8 + i
        estado = 'Aprobado' if ev.puntuacion >= 60 else 'Reprobado'
        fill = VERDE if ev.puntuacion >= 60 else ROJO
        valores = [i, f"Conversacion {ev.conversation_id}", ev.items_cumplidos,
                   ev.items_total, f"{round(ev.puntuacion)}%", estado]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill if col == 6 else GRIS
            cell.alignment = Alignment(horizontal='center')

    if evaluations:
        dr = 8 + len(evaluations) + 2
        ws.cell(row=dr, column=1, value='Detalle de criterios').font = Font(bold=True, size=12, color="1D4ED8")
        _fila_header(ws, ['Criterio', 'Cumplido', 'Keywords encontradas', 'Faltantes', 'Prohibidas'], dr+1)
        detalle = json.loads(evaluations[-1].detalle)
        for j, item in enumerate(detalle):
            r = dr + 2 + j
            ws.cell(row=r, column=1, value=item['criterio'])
            ws.cell(row=r, column=2, value='Si' if item['cumplido'] else 'No')
            ws.cell(row=r, column=3, value=', '.join(item['keywords_encontradas']))
            ws.cell(row=r, column=4, value=', '.join(item.get('keywords_faltantes', [])))
            ws.cell(row=r, column=5, value=', '.join(item.get('keywords_prohibidas', [])))
            fill2 = VERDE if item['cumplido'] else ROJO
            for col in range(1, 6):
                ws.cell(row=r, column=col).fill = fill2

    for col in ['A','B','C','D','E','F']:
        ws.column_dimensions[col].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def exportar_reporte_supervisor(mes: int, anio: int, db: Session) -> bytes:
    agents = db.query(Agent).filter(Agent.activo == True).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Supervisor"

    _cabecera(ws, 'Sistema de Supervision - Reporte General Supervisor', 'A1:H1', 1)
    ws['A3'] = 'Periodo:'; ws['A3'].font = BOLD_DARK; ws['B3'] = f"{MESES[mes]} {anio}"

    _fila_header(ws, ['Agente','Codigo','Area','Conversaciones','Promedio','Max','Min','Calificacion'], 5)

    row = 6
    todas_frecuencias = {}
    todas_faltantes = {}
    todas_prohibidas = {}

    for agent in agents:
        evaluations = db.query(Evaluation).filter(
            Evaluation.agent_id == agent.id,
            Evaluation.mes == mes,
            Evaluation.anio == anio
        ).all()
        if not evaluations:
            continue

        puntuaciones = [e.puntuacion for e in evaluations]
        promedio = round(sum(puntuaciones) / len(puntuaciones))

        if promedio >= 90:   calificacion = "Excelente"
        elif promedio >= 75: calificacion = "Bueno"
        elif promedio >= 60: calificacion = "Regular"
        else:                calificacion = "Deficiente"

        fill = VERDE if promedio >= 75 else (AMARILLO if promedio >= 60 else ROJO)

        valores = [agent.nombre, agent.codigo, agent.area, len(evaluations),
                   f"{promedio}%", f"{round(max(puntuaciones))}%",
                   f"{round(min(puntuaciones))}%", calificacion]
        for col, val in enumerate(valores, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = GRIS if col < 8 else fill
            cell.alignment = Alignment(horizontal='center')
        row += 1

        for ev in evaluations:
            for p, f in json.loads(ev.frecuencias or "{}").items():
                todas_frecuencias[p] = todas_frecuencias.get(p, 0) + f
            for p in json.loads(ev.palabras_obligatorias_faltantes or "[]"):
                todas_faltantes[p] = todas_faltantes.get(p, 0) + 1
            for p in json.loads(ev.palabras_prohibidas_encontradas or "[]"):
                todas_prohibidas[p] = todas_prohibidas.get(p, 0) + 1

    r = row + 2
    if todas_frecuencias:
        ws.cell(row=r, column=1, value='Palabras mas usadas').font = Font(bold=True, size=12, color="1D4ED8")
        _fila_header(ws, ['Palabra', 'Frecuencia'], r+1)
        for j, (p, f) in enumerate(sorted(todas_frecuencias.items(), key=lambda x: x[1], reverse=True)[:10]):
            ws.cell(row=r+2+j, column=1, value=p).fill = GRIS
            ws.cell(row=r+2+j, column=2, value=f).fill = GRIS

    if todas_faltantes:
        ws.cell(row=r, column=4, value='Obligatorias omitidas').font = Font(bold=True, size=12, color="DC2626")
        _fila_header(ws, ['Palabra', 'Veces omitida'], r+1)
        for j, (p, f) in enumerate(sorted(todas_faltantes.items(), key=lambda x: x[1], reverse=True)[:10]):
            ws.cell(row=r+2+j, column=4, value=p).fill = ROJO
            ws.cell(row=r+2+j, column=5, value=f).fill = ROJO

    if todas_prohibidas:
        ws.cell(row=r, column=7, value='Prohibidas detectadas').font = Font(bold=True, size=12, color="DC2626")
        _fila_header(ws, ['Palabra', 'Veces'], r+1)
        for j, (p, f) in enumerate(sorted(todas_prohibidas.items(), key=lambda x: x[1], reverse=True)):
            ws.cell(row=r+2+j, column=7, value=p).fill = ROJO
            ws.cell(row=r+2+j, column=8, value=f).fill = ROJO

    for col in ['A','B','C','D','E','F','G','H']:
        ws.column_dimensions[col].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def exportar_evaluacion_mensual(agent_id: int, mes: int, anio: int, db: Session) -> bytes:
    agent = db.query(Agent).filter(Agent.id == agent_id).first()
    ev = db.query(MonthlyEvaluation).filter(
        MonthlyEvaluation.agent_id == agent_id,
        MonthlyEvaluation.mes == mes,
        MonthlyEvaluation.anio == anio,
        MonthlyEvaluation.completada == True
    ).first()

    if not ev:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluacion Mensual"

    _cabecera(ws, 'Sistema de Supervision - Evaluacion Mensual de Desempeno', 'A1:D1', 1)

    ws['A3'] = 'Agente:';    ws['A3'].font = BOLD_DARK; ws['B3'] = agent.nombre if agent else 'N/A'
    ws['A4'] = 'Codigo:';    ws['A4'].font = BOLD_DARK; ws['B4'] = agent.codigo if agent else 'N/A'
    ws['A5'] = 'Periodo:';   ws['A5'].font = BOLD_DARK; ws['B5'] = f"{MESES[mes]} {anio}"
    ws['A6'] = 'Calificacion final:'; ws['A6'].font = BOLD_DARK; ws['B6'] = ev.calificacion_final
    ws['A7'] = 'Nivel:';     ws['A7'].font = BOLD_DARK; ws['B7'] = ev.nivel

    _fila_header(ws, ['Seccion', 'Criterio', 'Puntaje', 'Maximo'], 9)

    datos = [
        ('Automatico', 'Cumplimiento del checklist', ev.pts_cumplimiento, 20),
        ('Automatico', 'Uso de palabras obligatorias', ev.pts_obligatorias, 15),
        ('Automatico', 'Ausencia de palabras prohibidas', ev.pts_prohibidas, 15),
        ('Manual', 'Velocidad de respuesta', ev.velocidad_respuesta, 10),
        ('Manual', 'Capacidad de trabajo', ev.capacidad_trabajo, 10),
        ('Manual', 'Uso de plantillas', ev.uso_plantillas, 10),
        ('Manual', 'Resolucion en primer contacto', ev.resolucion_primer_contacto, 10),
        ('Manual', 'Actitud y trato al cliente', ev.actitud_trato, 10),
        ('Manual', 'Conocimiento del producto', ev.conocimiento_producto, 10),
    ]

    for i, (seccion, criterio, puntaje, maximo) in enumerate(datos):
        row = 10 + i
        fill = VERDE if seccion == 'Automatico' else PatternFill("solid", fgColor="EDE9FE")
        ws.cell(row=row, column=1, value=seccion).fill = fill
        ws.cell(row=row, column=2, value=criterio).fill = GRIS
        ws.cell(row=row, column=3, value=puntaje).fill = GRIS
        ws.cell(row=row, column=4, value=maximo).fill = GRIS

    if ev.observaciones:
        obs_row = 10 + len(datos) + 2
        ws.cell(row=obs_row, column=1, value='Observaciones:').font = BOLD_DARK
        ws.cell(row=obs_row+1, column=1, value=ev.observaciones)

    for col in ['A','B','C','D']:
        ws.column_dimensions[col].width = 30

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()